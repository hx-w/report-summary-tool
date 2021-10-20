# -*- coding: utf-8 -*-

import shutil
import datetime
import pandas as pd
import openpyxl

from typedef import *

align = openpyxl.styles.Alignment(
    horizontal='left', vertical='center', wrap_text=False)


def get_swsg_week_list(group_name: str) -> list:
    source_dir = os.path.join('.', group_name)
    pattern = re.compile(global_config.week_pattern)
    week_list = list(filter(
        lambda x: re.match(pattern, x),
        os.listdir(source_dir)
    ))
    week_list.sort(reverse=True)
    return week_list


def get_swmg_week_list(reverse: bool = True, sw=False) -> list:
    source_dir = os.path.join('.', global_config.summary)
    pattern = re.compile(global_config.week_pattern)
    week_list = list(filter(
        lambda x: re.match(pattern, x),
        os.listdir(source_dir)
    ))
    week_list.sort(reverse=reverse)
    if sw:
        return week_list
    week_list = list(filter(
        lambda x: os.path.exists(os.path.join(
            os.path.join(source_dir, x),
            f'{global_config.prefix}项目工作周报-{x}.xlsx'
        )),
        week_list
    ))
    return week_list


def get_week_count(start_week: str, end_week: str) -> int:
    week_list = get_swmg_week_list(False)
    return week_list.index(end_week) - week_list.index(start_week) + 1


def judge_swsg_result_exist(group_name: str, week: str) -> bool:
    source_dir = os.path.join('.', global_config.summary)
    result_file = os.path.join(
        source_dir,
        os.path.join(
            week, f'{global_config.prefix}小组工作周报-{week}-{group_name[3:]}.xlsx')
    )
    return os.path.exists(result_file)


def judge_swmg_result_exist(week: str) -> bool:
    source_dir = os.path.join('.', global_config.summary)
    result_file = os.path.join(
        source_dir,
        os.path.join(week, f'{global_config.prefix}项目工作周报-{week}.xlsx')
    )
    return os.path.exists(result_file)


def judge_mw_result_exist(start_week: str, end_week: str) -> bool:
    week_list = get_swmg_week_list(False)
    dist_path = os.path.join(
        global_config.summary,
        f'{global_config.prefix}项目工作报告-{"".join(start_week.split("-")[:2])}-{"".join(end_week.split("-")[::2])}.xlsx'
    )
    if start_week == week_list[0] and end_week == week_list[-1]:
        dist_path = os.path.join(
            global_config.summary, f'{global_config.prefix}项目工作报告.xlsx')
    return os.path.exists(dist_path)


def judge_swmg_group_complete(week: str) -> bool:
    trim_group = list(map(lambda x: x[3:], global_config.group))
    trim_group.sort()
    pattern = re.compile(f'{global_config.prefix}小组工作周报-{week}-(.*组).xlsx')
    source_dir = os.path.join('.', os.path.join(global_config.summary, week))
    group_list = list(filter(
        lambda x: re.match(pattern, x),
        os.listdir(source_dir)
    ))
    group_list = list(map(
        lambda x: re.findall(pattern, x)[0],
        group_list
    ))
    group_list.sort()
    return trim_group == group_list


def exec_swsg_merge(group_name: str, week: str, backup: bool = False) -> str:
    source_dir = os.path.join(
        '.', os.path.join(group_name, week)
    )
    pattern = re.compile(f'{global_config.prefix}个人工作周报-{week}-(.*).xlsx')
    filelist = list(filter(
        lambda x: re.match(pattern, x),
        os.listdir(source_dir)
    ))
    dist_dir = os.path.join(global_config.summary, week)
    if not os.path.exists(dist_dir):
        logger.debug(f'目录 "{dist_dir}" 不存在，已创建')
        os.makedirs(dist_dir)
    dist_path = os.path.join(
        dist_dir,
        f'{global_config.prefix}小组工作周报-{week}-{group_name[3:]}.xlsx'
    )
    # backup
    if backup and os.path.exists(dist_path):
        os.rename(dist_path, dist_path + '.bak')
    # cp template
    shutil.copyfile(global_config.template_group, dist_path)

    dist_book = openpyxl.load_workbook(dist_path)
    dist_sheet = dist_book['工作任务项']
    dist_sheet.delete_rows(2, dist_sheet.max_row + 1)
    col_name, col_date = 0, 0
    for idx in range(1, 1 + dist_sheet.max_column):
        if dist_sheet.cell(1, idx).value == '姓名':
            col_name = idx - 1
        elif dist_sheet.cell(1, idx).value == '日期（年/月/日）':
            col_date = idx - 1

    total_order = 1
    for eachfile in filelist:
        name = re.findall(pattern, eachfile)[0]
        full_path = os.path.join(source_dir, eachfile)
        source_sheet = openpyxl.load_workbook(
            full_path, read_only=True)['工作任务项']
        # print(source_sheet[1])
        for row_idx in range(2, source_sheet.max_row + 1):
            row_element = []
            for col_idx in range(1, source_sheet.max_column + 1):
                row_element.append(source_sheet.cell(
                    row=row_idx, column=col_idx).value)
            row_element.insert(col_name, name)
            row_element[col_date] = datetime.datetime.strftime(
                row_element[col_date], "%Y/%m/%d")
            row_element[0] = total_order
            dist_sheet.append(row_element)
            total_order += 1
    
    # align to left
    for row in dist_sheet.iter_rows(min_row=2):
        for col in row:
            col.alignment = align

    # update pivot table
    for sheet_idx in range(len(dist_book.sheetnames)):
        if dist_book.sheetnames[sheet_idx] == '工作任务项':
            continue
        pivot_sheet = dist_book[dist_book.sheetnames[sheet_idx]]
        pivot = pivot_sheet._pivots[0]  # 任何一个都可以共享同一个缓存
        boundary = f'A1:{chr(ord("A") + dist_sheet.max_column - 1)}{dist_sheet.max_row}'
        pivot.cache.cacheSource.worksheetSource.ref = boundary
        pivot.cache.refreshOnLoad = True  # 刷新加载

    dist_book.save(dist_path)

    logger.info(f'已经合并到文件 "{dist_path}"')
    return dist_path


def exec_swmg_merge(week: str, backup: bool = False) -> str:
    dist_dir = os.path.join(global_config.summary, week)
    dist_path = os.path.join(
        dist_dir,
        f'{global_config.prefix}项目工作周报-{week}.xlsx'
    )
    pattern = re.compile(f'{global_config.prefix}小组工作周报-{week}-(.*组).xlsx')
    filelist = list(filter(
        lambda x: re.match(pattern, x),
        os.listdir(dist_dir)
    ))

    if backup and os.path.exists(dist_path):
        os.rename(dist_path, dist_path + '.bak')

    # cp template
    shutil.copyfile(global_config.template_project, dist_path)

    dist_book = openpyxl.load_workbook(dist_path)
    dist_sheet = dist_book['工作任务项']
    dist_sheet.delete_rows(2, dist_sheet.max_row + 1)
    col_group, col_date = 0, 0
    for idx in range(1, 1 + dist_sheet.max_column):
        if dist_sheet.cell(1, idx).value == '组名':
            col_group = idx - 1
        elif dist_sheet.cell(1, idx).value == '日期（年/月/日）':
            col_date = idx - 1

    total_order = 1
    for eachfile in filelist:
        group_name = re.findall(pattern, eachfile)[0]
        full_path = os.path.join(dist_dir, eachfile)
        source_sheet = openpyxl.load_workbook(
            full_path, read_only=True)['工作任务项']
        # print(source_sheet[1])
        for row_idx in range(2, source_sheet.max_row + 1):
            row_element = []
            for col_idx in range(1, source_sheet.max_column + 1):
                row_element.append(source_sheet.cell(
                    row=row_idx, column=col_idx).value)
            row_element.insert(col_group, group_name)
            if not isinstance(row_element[col_date], str):
                row_element[col_date] = datetime.datetime.strftime(
                    row_element[col_date], "%Y/%m/%d")
            row_element[0] = total_order
            dist_sheet.append(row_element)
            total_order += 1
    
    # align to left
    for row in dist_sheet.iter_rows(min_row=2):
        for col in row:
            col.alignment = align

    # update pivot table
    for sheet_idx in range(len(dist_book.sheetnames)):
        if dist_book.sheetnames[sheet_idx] == '工作任务项':
            continue
        pivot_sheet = dist_book[dist_book.sheetnames[sheet_idx]]
        pivot = pivot_sheet._pivots[0]  # 任何一个都可以共享同一个缓存
        boundary = f'A1:{chr(ord("A") + dist_sheet.max_column - 1)}{dist_sheet.max_row}'
        pivot.cache.cacheSource.worksheetSource.ref = boundary
        pivot.cache.refreshOnLoad = True  # 刷新加载

    dist_book.save(dist_path)
    logger.info(f'已经合并到文件 "{dist_path}"')
    return dist_path


def exec_mw_merge(start_week: str, end_week: str, backup: bool = False) -> str:
    week_list = get_swmg_week_list(False)
    start_idx = week_list.index(start_week)
    end_idx = week_list.index(end_week)

    dist_path = os.path.join(
        global_config.summary,
        f'{global_config.prefix}项目工作报告-{"".join(start_week.split("-")[:2])}-{"".join(end_week.split("-")[::2])}.xlsx'
    )
    if start_idx == 0 and end_idx == len(week_list) - 1:
        dist_path = os.path.join(
            global_config.summary, f'{global_config.prefix}项目工作报告.xlsx')

    if backup and os.path.exists(dist_path):
        os.rename(dist_path, dist_path + '.bak')
    

    # cp template
    shutil.copyfile(global_config.template_project, dist_path)

    dist_book = openpyxl.load_workbook(dist_path)
    dist_sheet = dist_book['工作任务项']
    dist_sheet.delete_rows(2, dist_sheet.max_row + 1)
    col_date = 0
    for idx in range(1, 1 + dist_sheet.max_column):
        if dist_sheet.cell(1, idx).value == '日期（年/月/日）':
            col_date = idx - 1
            break
    
    total_order = 1
    for eachweek_idx in range(start_idx, end_idx + 1):
        eachweek = week_list[eachweek_idx]
        file_path = os.path.join(global_config.summary, os.path.join(
            eachweek, f'{global_config.prefix}项目工作周报-{eachweek}.xlsx'
        ))
        source_sheet = openpyxl.load_workbook(
            file_path, read_only=True)['工作任务项']
        # print(source_sheet[1])
        for row_idx in range(2, source_sheet.max_row + 1):
            row_element = []
            for col_idx in range(1, source_sheet.max_column + 1):
                row_element.append(source_sheet.cell(
                    row=row_idx, column=col_idx).value)
            if not isinstance(row_element[col_date], str):
                row_element[col_date] = datetime.datetime.strftime(
                    row_element[col_date], "%Y/%m/%d")
            row_element[0] = total_order
            dist_sheet.append(row_element)
            total_order += 1
    
    # align to left
    for row in dist_sheet.iter_rows(min_row=2):
        for col in row:
            col.alignment = align

    # update pivot table
    for sheet_idx in range(len(dist_book.sheetnames)):
        if dist_book.sheetnames[sheet_idx] == '工作任务项':
            continue
        pivot_sheet = dist_book[dist_book.sheetnames[sheet_idx]]
        pivot = pivot_sheet._pivots[0]  # 任何一个都可以共享同一个缓存
        boundary = f'A1:{chr(ord("A") + dist_sheet.max_column - 1)}{dist_sheet.max_row}'
        pivot.cache.cacheSource.worksheetSource.ref = boundary
        pivot.cache.refreshOnLoad = True  # 刷新加载

    dist_book.save(dist_path)
    # df_list = []
    # for eachweek_idx in range(start_idx, end_idx + 1):
    #     eachweek = week_list[eachweek_idx]
    #     file_path = os.path.join(global_config.summary, os.path.join(
    #         eachweek, f'{global_config.prefix}项目工作周报-{eachweek}.xlsx'
    #     ))
    #     df = pd.read_excel(file_path)
    #     df_list.append(df)

    # new_df = pd.concat(df_list)
    # new_df.reset_index(drop=True, inplace=True)
    # new_df['序号'] = pd.Series(list(range(1, len(new_df) + 1)))
    # new_df.to_excel(dist_path, '工作任务项', index=False)
    logger.info(f'已经合并到文件 "{dist_path}"')
    return dist_path


def judge_swsg_empty(group_name: str, week: str) -> bool:
    source_dir = os.path.join(
        '.', os.path.join(group_name, week)
    )
    pattern = re.compile(f'{global_config.prefix}个人工作周报-{week}-(.*).xlsx')
    filelist = list(filter(
        lambda x: re.match(pattern, x),
        os.listdir(source_dir)
    ))
    return len(filelist) == 0


def judge_swmg_empty(week: str) -> bool:
    source_dir = os.path.join(
        os.path.join('.', global_config.summary), week
    )
    # print(os.listdir(source_dir))
    return not (os.path.exists(source_dir) and len(os.listdir(source_dir)) > 0)


def get_swsg_preview(group_name: str, week: str) -> str:
    source_dir = os.path.join(
        '.', os.path.join(group_name, week)
    )
    pattern = re.compile(f'{global_config.prefix}个人工作周报-{week}-(.*).xlsx')
    filelist = list(filter(
        lambda x: re.match(pattern, x),
        os.listdir(source_dir)
    ))
    filelist = list(map(
        lambda x: re.findall(pattern, x)[0],
        filelist
    ))
    info_list = []
    for idx in range(0, len(filelist), 3):
        info_list.append((' ' * 4).join(filelist[idx:idx + 3]))
    return '\n\n'.join(info_list)


def get_swmg_preview(week: str) -> str:
    source_dir = os.path.join(
        os.path.join('.', global_config.summary), week
    )
    pattern = re.compile(f'{global_config.prefix}小组工作周报-{week}-(.*组).xlsx')
    filelist = list(filter(
        lambda x: re.match(pattern, x),
        os.listdir(source_dir)
    ))
    filelist = list(map(
        lambda x: re.findall(pattern, x)[0],
        filelist
    ))
    info_list = []
    for idx in range(0, len(filelist), 2):
        info_list.append((' ' * 3).join(filelist[idx:idx + 2]))
    return '\n\n'.join(info_list)
